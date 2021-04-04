import scipy as sp
import openpyxl as ox
import numpy as np
import matplotlib.pyplot as plt
from scipy.optimize import fsolve

# читаем данные

wb = ox.load_workbook(filename = './data.xlsx')
x = list()
y = list()
ws = wb['Sheet1']
for cell in  ws['A']:
	x.append(cell.value)
for cell in  ws['B']:
	y.append(cell.value)

# настраиваем детали отрисовки графиков
plt.figure(figsize=(8, 6))

plt.xlabel("T, K")
plt.ylabel("P, mmHg")

plt.autoscale(tight=True)

# рисуем исходные точки
plt.errorbar(x, y, xerr=0.5, yerr=0.05, fmt='.', ecolor='red')

legend = []

L = list()
for i in range(1, len(x)-1):
	L.append((y[i+1] - y[i-1])/(x[i+1] - x[i-1])*8.31*x[i]*x[i]/y[i])

avgL = 0
for i in range(len(L)):
	avgL += L[i]

avgL  = avgL/len(L)

sigma = 0;
for i in range(len(L)):
	sigma += (L[i] - avgL)*(L[i] - avgL)

sigma = sigma/len(L)
sigma = np.sqrt(sigma)

print("L = ")
print(avgL)

print("Погрешность = ")
print(sigma)
	

    # рисуем график модельной функции
#plt.plot(fx, f(fx), linewidth=2)


plt.legend(legend, loc="upper left")
plt.grid()
plt.savefig('data.png', dpi=50)
plt.show()
