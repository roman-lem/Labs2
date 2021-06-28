import scipy as sp
import openpyxl as ox
import numpy as np
import matplotlib.pyplot as plt
from scipy.optimize import fsolve

# читаем данные

wb = ox.load_workbook(filename = './data.xlsx')
x = list()
y = list()
ws = wb['Sheet1']
for cell in  ws['A']:
	x.append(float(cell.value))
for cell in  ws['B']:
	y.append(float(cell.value))

# настраиваем детали отрисовки графиков
plt.figure(figsize=(8, 6))

plt.xlabel("l, см")
plt.ylabel("P, Па")

plt.autoscale(tight=True)

# рисуем исходные точки
plt.errorbar(x, y, yerr=2, xerr=0.5, fmt='.', ecolor='red', color='red')

# аргументы для построения графиков моделей: исходный интервал + 60 дней
fx = sp.linspace(x[0]-10, x[len(x)-1]+30, 100)

    # получаем параметры модели для полинома степени 1
fp, residuals, rank, sv, rcond = sp.polyfit(x, y, 1, full=True)

    # функция-полином, если её напечатать, то увидите математическое выражение
f = sp.poly1d(fp)
print(f)

    # рисуем график модельной функции
#plt.plot(fx, f(fx), linewidth=2)




avgx = 0
avgy = 0
avgx2 = 0
avgy2 = 0
avgxy = 0
for i in range(3):
	avgx += x[i]
	avgy += y[i]
	avgx2 += x[i] * x[i]
	avgy2 += y[i] * y[i]
	avgxy += x[i] * y[i]

avgx = avgx/6
avgy = avgy/6
avgx2 = avgx2/6
avgy2 = avgy2/6
avgxy = avgxy/6

koef = (avgxy - avgx*avgy)/(avgx2 - avgx*avgx)

sigma = np.sqrt((avgy2 - avgy*avgy)/(avgx2 - avgx*avgx) - koef*koef)/np.sqrt(6)

print("Коэффициент = ")
print(koef)

print("Погрешность = ")
print(sigma)



plt.grid()
plt.savefig('data.png', dpi=50)
plt.show()
